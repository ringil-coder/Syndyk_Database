"""
Generuje plik Excel z danych w bazie SQLite (obwieszczenia.db).

Wymagania:
    pip install openpyxl

Użycie:
    python generate_excel.py
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

DB_FILE = Path(__file__).parent / "obwieszczenia.db"
OUTPUT_FILE = Path(__file__).parent / "obwieszczenia_masa_upadlosci.xlsx"

# Kolumny w Excelu (kolejność) -> kolumny w SQLite
COLUMNS = [
    ("Numer obwieszczenia", "numer_obwieszczenia"),
    ("Data obwieszczenia",  "data_obwieszczenia"),
    ("Sygnatura",           "sygnatura"),
    ("Imię",                "imie"),
    ("Nazwisko",            "nazwisko"),
    ("Miejsce zamieszkania","miejsce_zamieszkania"),
    ("Rodzaj podmiotu",     "rodzaj_podmiotu"),
    ("PESEL",               "pesel"),
    ("NIP",                 "nip"),
    ("Sąd",                 "sad"),
    ("Wydział",             "wydzial"),
    ("Treść obwieszczenia", "tresc_obwieszczenia"),
    ("Link",                "link"),
]


def read_from_db(db_path: Path = DB_FILE) -> list[dict]:
    if not db_path.exists():
        print(f"Baza danych {db_path} nie istnieje. Uruchom najpierw krz_scraper.py.")
        return []
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM obwieszczenia ORDER BY data_obwieszczenia DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def generate_excel(records: list[dict], output: Path = OUTPUT_FILE) -> None:
    if not records:
        print("Brak danych do zapisania.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Obwieszczenia"

    # Nagłówki
    header_names = [col[0] for col in COLUMNS]
    ws.append(header_names)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    link_font = Font(color="0563C1", underline="single")

    for rec in records:
        row_values = [rec.get(db_col, "") or "" for _, db_col in COLUMNS]
        ws.append(row_values)
        r = ws.max_row

        # Kolumna "Link" — wstaw hiperlink
        link_col_idx = header_names.index("Link") + 1
        link_val = rec.get("link", "")
        if link_val:
            cell = ws.cell(row=r, column=link_col_idx)
            cell.hyperlink = link_val
            cell.font = link_font

    # Auto-szerokość kolumn
    for col_idx, (name, _) in enumerate(COLUMNS, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            (len(str(ws.cell(row=i, column=col_idx).value or ""))
             for i in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output)
    print(f"Zapisano {len(records)} wierszy do {output}")


def main() -> None:
    records = read_from_db()
    generate_excel(records)


if __name__ == "__main__":
    main()

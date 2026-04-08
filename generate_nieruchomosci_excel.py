"""
Generuje plik Excel z danych w bazie SQLite (nieruchomosci.db).

Wymagania:
    pip install openpyxl

Uzycie:
    python generate_nieruchomosci_excel.py
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter

DB_FILE = Path(__file__).parent / "nieruchomosci.db"
OUTPUT_FILE = Path(__file__).parent / "nieruchomosci_syndyk.xlsx"

# Kolumny w Excelu (kolejnosc) -> kolumny w SQLite
COLUMNS = [
    ("Portal",            "portal"),
    ("Typ",               "typ_nieruchomosci"),
    ("Tytul",             "tytul"),
    ("Cena (PLN)",        "cena"),
    ("Pow. (m2)",         "powierzchnia_m2"),
    ("Cena/m2",           "cena_za_m2"),
    ("Pokoje",            "liczba_pokoi"),
    ("Pietro",            "pietro"),
    ("Miasto",            "miasto"),
    ("Dzielnica",         "dzielnica"),
    ("Ulica",             "ulica"),
    ("Rok budowy",        "rok_budowy"),
    ("Typ budynku",       "typ_budynku"),
    ("Stan",              "stan_wykonczenia"),
    ("Wlasnosc",          "forma_wlasnosci"),
    ("Rynek",             "rynek"),
    ("Waluta",            "waluta"),
    ("Ogloszeniodawca",   "ogloszeniodawca"),
    ("Data dodania",      "data_dodania"),
    ("Data scrape",       "data_scrape"),
    ("Opis",              "opis"),
    ("Link",              "url"),
]


def read_from_db(db_path: Path = DB_FILE) -> list[dict]:
    """Odczytuje dane z bazy SQLite."""
    if not db_path.exists():
        print(f"Baza danych {db_path} nie istnieje. Uruchom najpierw real_estate_scraper.py.")
        return []
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM nieruchomosci WHERE aktywne = 1 ORDER BY cena ASC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def generate_excel(records: list[dict], output: Path = OUTPUT_FILE) -> None:
    """Generuje plik Excel z danymi nieruchomosci."""
    if not records:
        print("Brak danych do zapisania.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Nieruchomosci"

    header_names = [col[0] for col in COLUMNS]

    # --- Naglowki ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")

    ws.append(header_names)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Zamroz naglowek
    ws.freeze_panes = "A2"

    link_font = Font(color="0563C1", underline="single")

    # --- Dane ---
    for rec in records:
        row_values = []
        for _, db_col in COLUMNS:
            val = rec.get(db_col, "") or ""
            # Skroc opis do 500 znakow w Excelu
            if db_col == "opis" and isinstance(val, str) and len(val) > 500:
                val = val[:500] + "..."
            row_values.append(val)
        ws.append(row_values)
        r = ws.max_row

        # Kolumna "Link" - wstaw hiperlink
        link_col_idx = header_names.index("Link") + 1
        link_val = rec.get("url", "")
        if link_val:
            cell = ws.cell(row=r, column=link_col_idx)
            cell.hyperlink = link_val
            cell.value = link_val
            cell.font = link_font

        # Formatowanie numeryczne
        cena_col = header_names.index("Cena (PLN)") + 1
        cena_m2_col = header_names.index("Cena/m2") + 1
        pow_col = header_names.index("Pow. (m2)") + 1

        for col_idx in [cena_col, cena_m2_col]:
            cell = ws.cell(row=r, column=col_idx)
            if cell.value and cell.value != "":
                try:
                    cell.value = float(cell.value)
                    cell.number_format = '#,##0'
                except (ValueError, TypeError):
                    pass

        pow_cell = ws.cell(row=r, column=pow_col)
        if pow_cell.value and pow_cell.value != "":
            try:
                pow_cell.value = float(pow_cell.value)
                pow_cell.number_format = '#,##0.0'
            except (ValueError, TypeError):
                pass

    # --- Auto-filter ---
    ws.auto_filter.ref = ws.dimensions

    # --- Auto-szerokosc kolumn ---
    max_widths = {
        "Tytul": 45,
        "Opis": 50,
        "Link": 40,
    }
    for col_idx, (name, _) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_limit = max_widths.get(name, 25)

        max_len = max(
            (len(str(ws.cell(row=i, column=col_idx).value or ""))
             for i in range(1, min(ws.max_row + 1, 102))),  # probka 100 wierszy
            default=len(name),
        )
        width = min(max(max_len + 2, len(name) + 2), max_limit)
        ws.column_dimensions[col_letter].width = width

    wb.save(output)
    print(f"Zapisano {len(records)} wierszy do {output}")


def main() -> None:
    records = read_from_db()
    generate_excel(records)


if __name__ == "__main__":
    main()

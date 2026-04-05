"""
Skrypt scrapujący obwieszczenia dotyczące masy upadłości
ze strony https://krz.ms.gov.pl/

Wymagania:
    pip install selenium openpyxl webdriver-manager

    (webdriver-manager automatycznie pobiera ChromeDriver pasujący do
    zainstalowanej wersji Chrome, dzięki czemu nie ma znaczenia jaki
    chromedriver znajduje się w PATH.)

Uwaga:
    Strona https://krz.ms.gov.pl/ jest aplikacją Angular, więc do pobrania
    danych konieczny jest Selenium (sama biblioteka requests nie wystarczy).
    Selektory CSS zawierają dynamiczne identyfikatory Angulara (ng-tns-c15-XX),
    dlatego w miarę możliwości używamy zapytań niezależnych od tych numerów.
"""

from __future__ import annotations

import time
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


BASE_URL = "https://krz.ms.gov.pl/"
OUTPUT_FILE = Path(__file__).parent / "obwieszczenia_masa_upadlosci.xlsx"
DEFAULT_WAIT = 20


def build_driver(headless: bool = True) -> webdriver.Chrome:
    """Tworzy sterownik Chrome, automatycznie pobierając pasujący
    ChromeDriver (ignorując wersję z PATH)."""
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--lang=pl-PL")

    # 1) Najpierw próbujemy webdriver-manager (pobiera sterownik
    #    pasujący do zainstalowanego Chrome i omija ten w PATH).
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        service = ChromeService(ChromeDriverManager().install())
        return webdriver.Chrome(service=service, options=opts)
    except Exception as exc:
        print(f"[info] webdriver-manager niedostępny ({exc}); "
              f"próbuję Selenium Manager.")

    # 2) Fallback: Selenium Manager (wbudowany w selenium >= 4.6).
    #    Pusty Service wymusza użycie Selenium Managera zamiast
    #    chromedrivera znalezionego w PATH.
    return webdriver.Chrome(service=ChromeService(), options=opts)


def safe_click(driver, element) -> None:
    """Klika element; jeśli jest przesłonięty, używa JS."""
    try:
        element.click()
    except Exception:
        driver.execute_script("arguments[0].click();", element)


def set_date_range_last_month(driver) -> None:
    """Ustawia zakres dat: ostatni miesiąc -> dziś."""
    wait = WebDriverWait(driver, DEFAULT_WAIT)

    today = date.today()
    month_ago = today - timedelta(days=14)
    fmt = "%d.%m.%Y"

    # Inputy dat (są wewnątrz panelu "Zakres dat publikacji") — pierwsze dwa
    # inputy typu text w panelu to data od i data do.
    date_inputs = wait.until(
        EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "p-calendar input.ui-inputtext")
        )
    )
    if len(date_inputs) < 2:
        raise RuntimeError("Nie znaleziono pól zakresu dat.")

    for inp, value in zip(date_inputs[:2], (month_ago.strftime(fmt), today.strftime(fmt))):
        inp.click()
        inp.send_keys(Keys.CONTROL, "a")
        inp.send_keys(Keys.DELETE)
        inp.send_keys(value)
        inp.send_keys(Keys.ESCAPE)
        time.sleep(0.3)


def expand_panel_by_header(driver, header_text: str) -> None:
    """Rozwija panel p-panel po tekście nagłówka (jeśli nie jest rozwinięty)."""
    header = driver.find_element(
        By.XPATH,
        f"//p-panel//span[contains(@class,'ui-panel-title') and "
        f"contains(normalize-space(.), \"{header_text}\")]",
    )
    toggler = header.find_element(
        By.XPATH,
        "./ancestor::div[contains(@class,'ui-panel-titlebar')]"
        "//a[contains(@class,'ui-panel-titlebar-icon')]",
    )
    # rozwiń tylko jeżeli jest zwinięty (ikona ma "ui-icon-plusthick")
    cls = toggler.find_element(By.TAG_NAME, "span").get_attribute("class") or ""
    if "plus" in cls:
        safe_click(driver, toggler)
        time.sleep(0.5)


def collapse_panel_by_header(driver, header_text: str) -> None:
    """Zwija panel p-panel po tekście nagłówka (jeśli jest rozwinięty)."""
    header = driver.find_element(
        By.XPATH,
        f"//p-panel//span[contains(@class,'ui-panel-title') and "
        f"contains(normalize-space(.), \"{header_text}\")]",
    )
    toggler = header.find_element(
        By.XPATH,
        "./ancestor::div[contains(@class,'ui-panel-titlebar')]"
        "//a[contains(@class,'ui-panel-titlebar-icon')]",
    )
    cls = toggler.find_element(By.TAG_NAME, "span").get_attribute("class") or ""
    if "minus" in cls:
        safe_click(driver, toggler)
        time.sleep(0.5)


def scrape() -> list[list]:
    driver = build_driver(headless=False)
    try:
        wait = WebDriverWait(driver, DEFAULT_WAIT)

        # 1) Strona główna — Angular boot może trwać dłużej
        driver.get(BASE_URL)
        print(f"[info] URL: {driver.current_url}")
        # Poczekaj aż Angular wyrenderuje menu
        for _ in range(30):
            if driver.find_elements(By.CSS_SELECTOR, "[id^='item-']"):
                break
            time.sleep(1)
        time.sleep(2)

        # 2) Menu -> "Tablica obwieszczeń".
        # Szukamy po tekście (najbardziej odporne) oraz po id=item-4.
        menu_item = None
        candidates = driver.find_elements(
            By.XPATH,
            "//*[contains(normalize-space(.), 'Tablica obwieszczeń')"
            " and (self::a or self::button or contains(@class,'menu'))]",
        )
        if candidates:
            # wybierz najgłębszy kliknięty element
            menu_item = candidates[-1]
        else:
            els = driver.find_elements(By.ID, "item-4")
            if els:
                menu_item = els[0]

        if menu_item is None:
            print("[warn] Nie znaleziono elementu menu 'Tablica obwieszczeń'.")
            print(f"[debug] Zawartość body (pierwsze 2000 znaków):\n"
                  f"{driver.find_element(By.TAG_NAME, 'body').text[:2000]}")
            raise RuntimeError("Brak elementu menu.")

        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", menu_item
        )
        time.sleep(0.5)
        # Próbujemy kliknąć kolejno: element, jego link <a>, rodzica
        clicked = False
        for target in (
            menu_item,
            *menu_item.find_elements(By.TAG_NAME, "a"),
            *menu_item.find_elements(By.XPATH, "./ancestor::a[1]"),
        ):
            try:
                driver.execute_script("arguments[0].click();", target)
                clicked = True
                break
            except Exception:
                continue
        if not clicked:
            raise RuntimeError("Nie udało się kliknąć w menu.")

        time.sleep(3)
        print(f"[info] Po kliknięciu URL: {driver.current_url}")

        # KRZ ładuje Tablicę obwieszczeń w iframe. Spróbuj znaleźć iframe
        # i przełączyć kontekst WebDrivera do niego.
        def switch_into_form_iframe() -> bool:
            driver.switch_to.default_content()
            iframes = driver.find_elements(By.TAG_NAME, "iframe")
            print(f"[info] Znaleziono {len(iframes)} iframe(ów).")
            for idx, f in enumerate(iframes):
                src = f.get_attribute("src") or ""
                print(f"[info] iframe[{idx}] src={src[:120]}")
            for f in iframes:
                try:
                    driver.switch_to.frame(f)
                    # Czy to iframe z formularzem?
                    if driver.find_elements(
                        By.CSS_SELECTOR,
                        "app-wyszukiwanie-obwieszczen-view, "
                        "p-calendar, "
                        "div.dodatkoweParametry",
                    ):
                        print("[info] Przełączono do iframe z formularzem.")
                        return True
                except Exception:
                    pass
                driver.switch_to.default_content()
            return False

        # Poczekaj aż iframe się załaduje i znajdź ten właściwy
        form_ready = False
        for attempt in range(15):
            if switch_into_form_iframe():
                form_ready = True
                break
            # może formularz jest jednak w głównym dokumencie
            driver.switch_to.default_content()
            if driver.find_elements(
                By.CSS_SELECTOR,
                "app-wyszukiwanie-obwieszczen-view, div.dodatkoweParametry",
            ):
                form_ready = True
                print("[info] Formularz w głównym dokumencie.")
                break
            time.sleep(1)

        if not form_ready:
            print("[debug] Body (pierwsze 2000 znaków):\n"
                  f"{driver.find_element(By.TAG_NAME, 'body').text[:2000]}")
            raise RuntimeError(
                "Nie znaleziono formularza wyszukiwania obwieszczeń."
            )
        time.sleep(2)

        # 3) Zakres dat: ostatni miesiąc
        set_date_range_last_month(driver)

        # 4) Rozwiń dodatkowe parametry
        extra = wait.until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "div.dodatkoweParametry")
            )
        )
        safe_click(driver, extra)
        time.sleep(1)

        # 5-7) Zwijanie odbywa się później, po rozwinięciu wszystkich
        # paneli w kroku 8.

        # 8) W kategoriach obwieszczeń wybierz TYLKO pozycję dotyczącą
        #    obwieszczeń o masie upadłości.
        # Najpierw rozwiń wszystkie zwinięte panele w sekcji "dodatkowe
        # parametry", żeby checkboxy (w tym "masa upadłości") były
        # widoczne i klikalne. Panel #ui-panel-9-label (kategorie
        # obwieszczeń) często jest domyślnie zwinięty.
        togglers = driver.find_elements(
            By.CSS_SELECTOR,
            "p-panel a.ui-panel-titlebar-icon",
        )
        print(f"[info] Znaleziono {len(togglers)} togglerów paneli.")
        for t in togglers:
            try:
                icon = t.find_element(By.TAG_NAME, "span")
                cls = icon.get_attribute("class") or ""
                if "plus" in cls:
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", t
                    )
                    driver.execute_script("arguments[0].click();", t)
                    time.sleep(0.3)
            except Exception:
                pass
        time.sleep(0.8)

        # Zwiń z powrotem 3 panele z postępowaniami, żeby ich checkboxy
        # nie zostały zaznaczone.
        for header in (
            "Postępowania restrukturyzacyjne",
            "Postępowania upadłościowe",
            "Postępowania w przedmiocie ogłoszenia upadłości",
        ):
            try:
                collapse_panel_by_header(driver, header)
            except Exception:
                pass
        time.sleep(0.5)

        # Odznacz wszystkie widoczne, zaznaczone checkboxy
        checkboxes = driver.find_elements(
            By.CSS_SELECTOR,
            "p-checkbox .ui-chkbox-box",
        )
        for cb in checkboxes:
            cls = cb.get_attribute("class") or ""
            if "ui-state-active" in cls and cb.is_displayed():
                try:
                    driver.execute_script("arguments[0].click();", cb)
                    time.sleep(0.05)
                except Exception:
                    pass

        # Zaznacz 9. pozycję w panelu kategorii obwieszczeń
        # (odpowiednik xpath //*[@id="ui-panel-9-content"]/div/div/div[9])
        target = None
        try:
            target = driver.find_element(
                By.XPATH,
                '//*[@id="ui-panel-9-content"]/div/div/div[9]',
            )
            print("[info] Znaleziono ui-panel-9-content/div/div/div[9].")
        except Exception as exc:
            print(f"[warn] Nie znaleziono bezpośredniego xpath: {exc}")
            # Diagnostyka: wypisz wszystkie panele -content
            for panel in driver.find_elements(
                By.CSS_SELECTOR, "[id^='ui-panel-'][id$='-content']"
            ):
                pid = panel.get_attribute("id")
                kids = panel.find_elements(By.XPATH, "./div/div/div")
                print(f"  - {pid}: {len(kids)} dzieci div/div/div")
            # Fallback — panel z >=9 dzieci zawierający "masy upadłości"
            for panel in driver.find_elements(
                By.CSS_SELECTOR, "[id^='ui-panel-'][id$='-content']"
            ):
                children = panel.find_elements(By.XPATH, "./div/div/div")
                if len(children) >= 9 and any(
                    "masy upadłości" in (c.text or "").lower()
                    for c in children
                ):
                    target = children[8]
                    break

        if target is not None:
            text_preview = target.text.strip()[:100]
            print(f"[info] Element docelowy: '{text_preview}'")
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", target
            )
            time.sleep(0.3)

            # Znajdź właściwy checkbox PrimeNG wewnątrz tego divu.
            click_candidates = (
                target.find_elements(By.CSS_SELECTOR, "p-checkbox .ui-chkbox-box")
                + target.find_elements(By.CSS_SELECTOR, ".ui-chkbox-box")
                + target.find_elements(By.CSS_SELECTOR, "p-checkbox")
                + target.find_elements(By.TAG_NAME, "label")
            )
            print(f"[info] Kandydaci do kliknięcia: {len(click_candidates)}")

            clicked_ok = False
            for cand in click_candidates:
                try:
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", cand
                    )
                    # użyj zdarzenia myszy (bardziej niezawodne dla PrimeNG)
                    driver.execute_script(
                        "var e=arguments[0];"
                        "['mousedown','mouseup','click'].forEach(function(t){"
                        "e.dispatchEvent(new MouseEvent(t,{bubbles:true,"
                        "cancelable:true,view:window}));"
                        "});",
                        cand,
                    )
                    time.sleep(0.3)
                    # sprawdź czy checkbox stał się aktywny
                    box = target.find_elements(
                        By.CSS_SELECTOR, ".ui-chkbox-box"
                    )
                    if box and "ui-state-active" in (
                        box[0].get_attribute("class") or ""
                    ):
                        clicked_ok = True
                        print("[info] Checkbox zaznaczony.")
                        break
                except Exception as exc:
                    print(f"[warn] Klik zawiódł: {exc}")
                    continue

            if not clicked_ok:
                # ostatnia próba — klik w sam div
                driver.execute_script("arguments[0].click();", target)
                time.sleep(0.3)
                box = target.find_elements(By.CSS_SELECTOR, ".ui-chkbox-box")
                if box and "ui-state-active" in (
                    box[0].get_attribute("class") or ""
                ):
                    print("[info] Checkbox zaznaczony (fallback).")
                else:
                    print("[warn] Nie udało się zaznaczyć checkboxa.")
        else:
            print("[warn] Nie znaleziono panelu z 9. pozycją 'masa upadłości'.")
        time.sleep(0.5)

        # 9) Kliknij przycisk "Szukaj" / "Wyszukaj"
        search_btn = None
        for xp in (
            "//button[.//span[contains(normalize-space(.),'Szukaj')]]",
            "//button[.//span[contains(normalize-space(.),'Wyszukaj')]]",
            "//button[contains(normalize-space(.),'Szukaj')]",
            "//button[contains(normalize-space(.),'Wyszukaj')]",
            "//button[contains(@class,'primary')]",
        ):
            els = driver.find_elements(By.XPATH, xp)
            els = [e for e in els if e.is_displayed()]
            if els:
                search_btn = els[0]
                print(f"[info] Przycisk szukaj znaleziony: xpath={xp}")
                break
        if search_btn is None:
            # Dump listę wszystkich widocznych przycisków
            btns = driver.find_elements(By.TAG_NAME, "button")
            print("[debug] Widoczne przyciski:")
            for b in btns:
                if b.is_displayed():
                    print(f"  - '{b.text.strip()}' class={b.get_attribute('class')}")
            raise RuntimeError("Nie znaleziono przycisku wyszukiwania.")
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", search_btn
        )
        safe_click(driver, search_btn)

        # 10) Rozwiń panel z wynikami (np. ui-panel-12 / ui-panel-13).
        # Najpierw poczekaj aż wyniki się załadują, potem rozwiń panel.
        time.sleep(3)

        def expand_results_panel() -> bool:
            # Znajdź panel(e) wyników po nagłówku zawierającym "Wyniki"
            # LUB "obwieszczeń" w tytule, albo dowolny p-panel zawierający
            # p-table. Rozwiń jeśli zwinięty.
            panels = driver.find_elements(By.CSS_SELECTOR, "p-panel")
            for p in panels:
                # sprawdź czy zawiera p-table (lub będzie zawierać gdy
                # zostanie rozwinięty — patrz title)
                has_table = bool(p.find_elements(By.CSS_SELECTOR, "p-table"))
                title_els = p.find_elements(
                    By.CSS_SELECTOR, ".ui-panel-title"
                )
                title = title_els[0].text.strip() if title_els else ""
                if has_table or "wynik" in title.lower() or \
                        "obwieszcze" in title.lower():
                    togglers = p.find_elements(
                        By.CSS_SELECTOR, "a.ui-panel-titlebar-icon"
                    )
                    if not togglers:
                        continue
                    tog = togglers[0]
                    icon = tog.find_element(By.TAG_NAME, "span")
                    cls = icon.get_attribute("class") or ""
                    if "plus" in cls:
                        driver.execute_script(
                            "arguments[0].scrollIntoView({block:'center'});",
                            tog,
                        )
                        driver.execute_script("arguments[0].click();", tog)
                        time.sleep(1)
                        print(f"[info] Rozwinięto panel: '{title}'")
                    else:
                        print(f"[info] Panel '{title}' już rozwinięty.")
                    return True
            return False

        expanded = False
        for _ in range(20):
            if expand_results_panel():
                expanded = True
                break
            time.sleep(1)

        if not expanded:
            print("[warn] Nie znaleziono panelu z wynikami; "
                  "próbuję ui-panel-12/13 bezpośrednio.")
            for pid in ("ui-panel-12", "ui-panel-13", "ui-panel-11"):
                tog = driver.find_elements(
                    By.CSS_SELECTOR,
                    f"#{pid} a.ui-panel-titlebar-icon",
                )
                if tog:
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});",
                        tog[0],
                    )
                    driver.execute_script("arguments[0].click();", tog[0])
                    time.sleep(1)
                    print(f"[info] Kliknąłem toggler {pid}.")
                    break
        time.sleep(2)

        # 11) Zbierz tabelę wyników
        wait.until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "p-table table")
            )
        )
        time.sleep(1)

        rows_data: list[list] = []
        # Nagłówki
        header_cells = driver.find_elements(
            By.CSS_SELECTOR, "p-table table thead th"
        )
        headers = [th.text.strip() for th in header_cells] or [
            "Lp.", "Data publikacji", "Sygnatura", "Kategoria",
            "Podmiot", "Szczegóły",
        ]

        # Iteracja po wszystkich stronach paginacji
        seen_page_signatures: set[str] = set()
        while True:
            body_rows = driver.find_elements(
                By.CSS_SELECTOR, "p-table table tbody tr"
            )
            page_signature = "|".join(r.text for r in body_rows)
            if page_signature in seen_page_signatures:
                break
            seen_page_signatures.add(page_signature)

            for tr in body_rows:
                cells = tr.find_elements(By.TAG_NAME, "td")
                row = []
                for td in cells:
                    links = td.find_elements(By.TAG_NAME, "a")
                    if links:
                        href = links[0].get_attribute("href") or ""
                        label = links[0].text.strip() or td.text.strip() or href
                        row.append({"text": label, "href": href})
                    else:
                        row.append(td.text.strip())
                rows_data.append(row)

            # Następna strona
            next_btns = driver.find_elements(
                By.CSS_SELECTOR,
                "p-paginator .ui-paginator-next:not(.ui-state-disabled)",
            )
            if not next_btns:
                break
            safe_click(driver, next_btns[0])
            time.sleep(1.5)

        return [headers, *rows_data]
    finally:
        driver.quit()


def fetch_details_for_links(rows_data: list[list]) -> None:
    """Dla każdego wiersza otwiera jego link i dokleja teksty z paneli
    #ui-panel-2, #ui-panel-3 i #ui-panel-4 do struktury wiersza."""
    # Zbierz unikalne URL-e do odwiedzenia
    urls: list[str] = []
    for row in rows_data:
        for cell in row:
            if isinstance(cell, dict) and cell.get("href"):
                urls.append(cell["href"])
                break  # pierwszy link w wierszu wystarczy
        else:
            urls.append("")

    if not any(urls):
        print("[info] Brak linków w tabeli — pomijam etap szczegółów.")
        return

    driver = build_driver(headless=False)
    try:
        wait = WebDriverWait(driver, 30)
        for idx, url in enumerate(urls, start=1):
            if not url:
                rows_data[idx - 1].extend(["", "", ""])
                continue
            print(f"[info] ({idx}/{len(urls)}) Pobieram szczegóły: {url}")
            driver.get(url)
            time.sleep(3)

            # KRZ ładuje szczegóły w iframe — przełącz się
            switched = False
            driver.switch_to.default_content()
            for f in driver.find_elements(By.TAG_NAME, "iframe"):
                try:
                    driver.switch_to.frame(f)
                    if driver.find_elements(
                        By.CSS_SELECTOR,
                        "#ui-panel-2, #ui-panel-3, #ui-panel-4, p-panel",
                    ):
                        switched = True
                        break
                except Exception:
                    pass
                driver.switch_to.default_content()

            # Poczekaj aż panele się załadują
            try:
                wait.until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR,
                         "#ui-panel-2, #ui-panel-3, #ui-panel-4")
                    )
                )
            except Exception:
                print(f"[warn] Panele nie załadowały się dla {url}")
                rows_data[idx - 1].extend(["", "", ""])
                continue
            time.sleep(1)

            texts = []
            for pid in ("ui-panel-2", "ui-panel-3", "ui-panel-4"):
                els = driver.find_elements(By.ID, pid)
                if els:
                    # upewnij się, że panel jest rozwinięty
                    togglers = els[0].find_elements(
                        By.CSS_SELECTOR, "a.ui-panel-titlebar-icon"
                    )
                    if togglers:
                        icon = togglers[0].find_elements(By.TAG_NAME, "span")
                        if icon and "plus" in (
                            icon[0].get_attribute("class") or ""
                        ):
                            driver.execute_script(
                                "arguments[0].click();", togglers[0]
                            )
                            time.sleep(0.5)
                    texts.append(els[0].text.strip())
                else:
                    texts.append("")
            rows_data[idx - 1].extend(texts)
    finally:
        driver.quit()


def scrape_with_details() -> tuple[list, list[list]]:
    data = scrape()
    headers, rows_data = data[0], list(data[1:])
    headers = list(headers) + ["Panel 2", "Panel 3", "Panel 4"]
    fetch_details_for_links(rows_data)
    # uzupełnij krótsze wiersze
    for row in rows_data:
        while len(row) < len(headers):
            row.append("")
    return headers, rows_data


def save_to_excel(data: list[list], path: Path) -> None:
    if not data:
        print("Brak danych do zapisania.")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Obwieszczenia"

    headers = data[0]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    link_font = Font(color="0563C1", underline="single")
    for row in data[1:]:
        ws.append([
            cell["text"] if isinstance(cell, dict) else cell for cell in row
        ])
        r = ws.max_row
        for col_idx, cell in enumerate(row, start=1):
            if isinstance(cell, dict) and cell.get("href"):
                xcell = ws.cell(row=r, column=col_idx)
                xcell.hyperlink = cell["href"]
                xcell.font = link_font

    for col_idx, _ in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            (len(str(ws.cell(row=i, column=col_idx).value or ""))
             for i in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(path)
    print(f"Zapisano {ws.max_row - 1} wierszy do {path}")


def main() -> None:
    headers, rows_data = scrape_with_details()
    save_to_excel([headers, *rows_data], OUTPUT_FILE)


if __name__ == "__main__":
    main()
